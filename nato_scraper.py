import os
import pickle
import re
import openpyxl
import pandas as pd
import pdftables_api
import requests
from collections import OrderedDict
from urllib.parse import urljoin
from bs4 import BeautifulSoup

import helpers


BASE_URL = 'http://www.nspa.nato.int/en/organization/procurement/contract.htm'
API_KEY = '1rxi0it7pla0'

# data frame column headers
DF_HEADERS = [
    'contractor',
    'country',
    'value_in_euro',
    'order_purpose',
    'period_start',
    'period_end',
    'year',
]


def extract_row_data(sheet):
    """
    Extract the data from a sheet
    :return: list of lists containing the data form each row
    """

    sheet_rows = [row for row in sheet.iter_rows()][4:-1]

    data_rows = list()
    for row in sheet_rows:
        row_values = list()
        for cell in row:
            row_values.append(cell.value)
        data_rows.append(row_values)

    del sheet_rows

    # If this sheet does not have 4 or 5 columns exit the function
    if len(data_rows[0]) < 4:
        return

    # If this sheet has 5 columns, concatenate the last two columns on each row
    if len(data_rows[0]) == 5:
        for index, row in enumerate(data_rows):
            if not len(row) == 5:
                continue

            if row[0] and (not row[1]) and (not row[2]) and isinstance(row[3], str) and isinstance(row[4], str):
                data_rows[index][3] = data_rows[index][3] + ' ' + data_rows[index][4]
                data_rows[index].pop(4)
                continue

            if (row[0] and isinstance(row[1], str) and (not row[2]) and
                    (not row[2]) and (not row[3]) and (len(row[1]) > 3)):
                data_rows[index][0] = data_rows[index][0] + data_rows[index][1]
                data_rows[index].pop(-1)
                continue

            if row[4] and not row[1]:
                data_rows[index].pop(1)
                continue

            if row[4] and isinstance(row[3], int) and isinstance(row[1], str):
                data_rows[index][0] = data_rows[index][0] + ' ' + data_rows[index][1]
                data_rows[index].pop(1)
                continue

            if row[1] and isinstance(row[3], str) and isinstance(row[4], str):
                data_rows[index][3] = data_rows[index][3] + ' ' + data_rows[index][4]
                data_rows[index].pop(4)
                continue

    # Check for errors resulted from the OCR process

    index = 0
    data_rows_len = len(data_rows)
    while index < data_rows_len:

        # If the current row is empty except for column D
        if (not data_rows[index][0]) and (not data_rows[index][1]) and (not data_rows[index][2]):

            # If the index+1 row is not empty on any of the columns and row index+2 is empty on all columns except
            # column D
            if (((index + 2) < data_rows_len) and ((index + 3) == data_rows_len) and
                    data_rows[index+1][0] and data_rows[index+1][1] and data_rows[index+1][2] and
                    data_rows[index+1][3] and (not data_rows[index+2][0]) and (not data_rows[index+2][1]) and
                    (not data_rows[index+2][2]) and data_rows[index+2][3]):
                data_rows[index][0] = data_rows[index + 1][0]
                data_rows[index][1] = data_rows[index + 1][1]
                data_rows[index][2] = data_rows[index + 1][2]
                data_rows[index][3] = str(data_rows[index][3]) + ' ' + str(data_rows[index + 1][3])
                data_rows.pop(index + 1)
                data_rows[index][3] = str(data_rows[index][3]) + ' ' + str(data_rows[index + 1][3])
                data_rows.pop(index + 1)
                data_rows_len = len(data_rows)
                index += 1
                continue

            # If row index+1 is empty only on column D and row index+2 is empty except on column D
            elif (((index + 2) < data_rows_len) and
                    data_rows[index+1][0] and data_rows[index+1][1] and data_rows[index+1][2] and
                    (not data_rows[index+1][3]) and (not data_rows[index+2][0]) and (not data_rows[index+2][1]) and
                    (not data_rows[index+2][2]) and data_rows[index+2][3]):

                data_rows[index][0] = data_rows[index + 1][0]
                data_rows[index][1] = data_rows[index + 1][1]
                data_rows[index][2] = data_rows[index + 1][2]
                data_rows[index][3] = str(data_rows[index][3]) + ' ' + str(data_rows[index + 2][3])
                data_rows.pop(index + 1)
                data_rows.pop(index + 1)
                data_rows_len = len(data_rows)
                index += 1
                continue

            elif (((index + 1) < data_rows_len) and
                    data_rows[index+1][0] and data_rows[index+1][1] and data_rows[index+1][2] and
                    data_rows[index+1][3]):

                data_rows[index][0] = data_rows[index + 1][0]
                data_rows[index][1] = data_rows[index + 1][1]
                data_rows[index][2] = data_rows[index + 1][2]
                data_rows[index][3] = str(data_rows[index][3]) + ' ' + str(data_rows[index + 1][3])
                data_rows.pop(index + 1)
                data_rows_len = len(data_rows)
                index += 1
                continue

        # If on the current row column A is non-empty and columns B and C are empty
        if data_rows[index][0] and not data_rows[index][1] and data_rows[index][2]:

            # If columnA and D on the current row is not empty and column A and D are empty on row index+1 and not
            # empty on row index+2 and column B and C are non-empty on row index+1 and empty on row index+2
            if (((index + 2) < data_rows_len) and data_rows[index][3] and (not data_rows[index+1][0]) and
                    (not data_rows[index+1][3]) and data_rows[index+1][1] and data_rows[index+1][2] and
                    data_rows[index+2][0] and data_rows[index+2][3] and (not data_rows[index+2][1]) and
                    (not data_rows[index+2][2])):

                data_rows[index][0] = data_rows[index][0] + ' ' + data_rows[index + 2][0]
                data_rows[index][1] = data_rows[index + 1][1]
                data_rows[index][2] = data_rows[index + 1][2]
                data_rows[index][3] = str(data_rows[index][3]) + ' ' + str(data_rows[index + 2][3])
                data_rows.pop(index + 1)
                data_rows.pop(index + 1)
                data_rows_len = len(data_rows)
                index += 1
                continue

            # If column A is empty on row index+1 and not empty on row index and index+1 and columns B through D
            # aren't empty on row index+1
            elif (((index + 2) < data_rows_len) and (not data_rows[index][3]) and (not data_rows[index+1][0]) and
                    data_rows[index+1][3] and data_rows[index+1][1] and data_rows[index+1][2] and
                    data_rows[index+2][0] and (not data_rows[index+2][3]) and (not data_rows[index+2][1]) and
                    (not data_rows[index+2][2])):

                data_rows[index][0] = data_rows[index][0] + ' ' + data_rows[index + 2][0]
                data_rows[index][1] = data_rows[index + 1][1]
                data_rows[index][2] = data_rows[index + 1][2]
                data_rows[index][3] = data_rows[index][3]
                data_rows.pop(index + 1)
                data_rows.pop(index + 1)
                data_rows_len = len(data_rows)
                index += 1
                continue

            # If on row index only column A is non empty and on row index+1 all columns are non-empty
            elif (((index + 1) < data_rows_len) and (not data_rows[index][1]) and (not data_rows[index][2]) and
                    (not data_rows[index][3]) and data_rows[index+1][0] and data_rows[index+1][1] and
                    data_rows[index+1][2] and data_rows[index+1][3]):
                data_rows[index][0] = data_rows[index][0] + ' ' + data_rows[index + 1][0]
                data_rows[index][1] = data_rows[index + 1][1]
                data_rows[index][2] = data_rows[index + 1][2]
                data_rows[index][3] = data_rows[index + 1][3]
                data_rows.pop(index + 1)
                data_rows_len = len(data_rows)
                index += 1
                continue

        index += 1
    return data_rows


def generate_pdf_files():
    """
    Download the NATO Support and Procurement Agency Contract Awards
    pdf lists and yield their names

    :return: Yield pdf file paths
    """

    response = helpers.request(BASE_URL)
    soup = BeautifulSoup(response, 'lxml')

    yearly_reports = dict()
    for year in soup.select('div.boxContent p.bold'):
        yearly_reports[year.get_text().strip()] = year.findNextSibling('ul').findAll('a')

    for year in yearly_reports:
        for report_link in yearly_reports[year]:

            name = report_link['href']
            name = 'pdfs/' + year + '__' + re.sub('/PDF/Procurement/', '', name)

            link = urljoin(BASE_URL, report_link['href'])
            pdf_response = requests.get(link)
            with open(name, 'wb') as pdf:
                for chunk in pdf_response.iter_content():
                    if chunk:
                        pdf.write(chunk)
            yield name


def generate_xlsx_files():
    """
    Convert all pdf files to xlsx files and extract the rows from them
    :return: Yield lists of rows containing the extracted data
    """

    # If there is no file containing the list of already converted pdf files create one
    # and add an empty list into it
    converted_list_path = os.path.join(os.path.curdir, 'excels/converted_list.txt')
    if not os.path.exists(converted_list_path):
        converted_list_file = open(converted_list_path, 'wb')
        pickle.dump(list(), converted_list_file)

    with open(converted_list_path, 'rb') as converted_list_file:
        converted_list = pickle.load(converted_list_file)

    # Start pdftables api client
    client = pdftables_api.Client(API_KEY)

    for pdf_name in generate_pdf_files():
        rows = list()
        # Get the name of the output csv file
        xlsx_name = re.sub('.*?/|\.pdf', '', pdf_name).strip()
        xlsx_name = 'excels/' + xlsx_name + '.xlsx'

        # If the current pdf is not in the list of already converted
        if pdf_name not in converted_list:
            client.xlsx(pdf_name, xlsx_name)
            converted_list.append(pdf_name)

        # Extract the start date, end date, and year of the period
        year = re.match('pdfs/([0-9]+)__.*\.pdf', pdf_name)
        if year:
            year = int(year.groups()[0])
        else:
            year = 0

        period_start = re.match('pdfs/[0-9]{4}__(.*?)-.*', pdf_name)
        if period_start:
            period_start = period_start.groups()[0].strip()
            period_start = period_start[-2:] + ' ' + period_start[:-2]
            period_start = period_start + ' ' + str(year)
            period_start = helpers.get_date_text(period_start, romanian=False)
        else:
            period_start = ''

        period_end = re.match('pdfs/[0-9]{4}__.*?-(.*?)\.pdf', pdf_name)
        if period_end:
            period_end = period_end.groups()[0].strip()
            period_end = period_end[-2:] + ' ' + period_end[:-2]
            period_end = period_end + ' ' + str(year)
            period_end = helpers.get_date_text(period_end, romanian=False)
        else:
            period_end = ''

        workbook = openpyxl.load_workbook(xlsx_name)

        # Initialize a variable with the page at which the extraction should stop
        page = workbook.worksheets[0].max_row
        if not workbook.worksheets[0][('C' + str(page))].value:
            if workbook.worksheets[0][('D' + str(page))].value:
                page = workbook.worksheets[0][('D' + str(page))].value
            elif workbook.worksheets[0][('B' + str(page))].value:
                page = workbook.worksheets[0][('B' + str(page))].value
        else:
            page = workbook.worksheets[0][('C' + str(page))].value


        max_page = re.sub('.*of\s*', '', page).strip()

        for sheet in workbook.worksheets:

            if sheet['B' + str(sheet.max_row)].value:
                page = sheet['B' + str(sheet.max_row)].value

            elif sheet['C' + str(sheet.max_row)].value:
                page = sheet['C' + str(sheet.max_row)].value

            elif sheet['D' + str(sheet.max_row)].value:
                page = sheet['D' + str(sheet.max_row)].value

            page = re.sub('of.*', '', page).strip()

            sheet_rows = extract_row_data(sheet)
            if not sheet_rows:
                break

            for index, row in enumerate(sheet_rows):
                sheet_rows[index].extend([period_start, period_end, year])

            rows.extend(sheet_rows)

            if page == max_page:
                break

        yield rows

    with open(converted_list_path, 'wb') as converted_list_file:
        pickle.dump(converted_list, converted_list_file)


def build_data_frame():
    """
    Build the data frame in memory and return it
    :return: data frame containing all the extracted data
    """
    df_rows = list()
    for rows in generate_xlsx_files():
        df_rows.extend(rows)

    data = OrderedDict()
    for index, key in enumerate(DF_HEADERS):
        data[key] = [row[index] for row in df_rows]

    indexes = [index for index in range(1, len(df_rows)+1)]
    data_frame = pd.DataFrame(data=data, index=indexes)
    return data_frame


def main():
    """
    Main function of the scraper. Handles the exectuion of the data frame building function
    and the output of the data to a csv file
    """
    nato_data = build_data_frame()
    nato_data.to_csv('nato_output.csv', sep=',')


if __name__ == '__main__':
    main()
